"""Export all payment_outs rows to Excel or CSV (e.g. in a OneDrive-synced folder)."""

from __future__ import annotations

import asyncio
import logging
import time
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from config import Settings
from database import (
    PaymentRecord,
    get_paidside_epoch,
    list_all_payments,
    list_payments_since,
)
from handlers.stats_period import stats_timezone
from money_format import format_amount as _format_amount

logger = logging.getLogger(__name__)

_EXCEL_SYNC_LOCK: asyncio.Lock | None = None
SYNC_ESTIMATE_SECONDS = 6
LOCAL_WRITE_RETRY_SECONDS = 30


def _export_output_path(settings: Settings) -> Path | None:
    """Return the Excel export path for this bot instance, or None if disabled."""
    if not settings.payments_onedrive_path:
        return None
    return Path(settings.payments_onedrive_path).expanduser().resolve()

HEADERS = (
    "Amount",
    "Date",
    "Starter",
    "Finisher",
    "Card",
    "Cleared",
    "Paying Starter",
    "Paying Finisher",
    "Paying Centre",
)

STARTER_PAY_RATE = 0.05
FINISHER_PAY_RATE = 0.15
CENTRE_PAY_RATE = 0.20

STARTER_PAY_PERCENT = int(STARTER_PAY_RATE * 100)
FINISHER_PAY_PERCENT = int(FINISHER_PAY_RATE * 100)
CENTRE_PAY_PERCENT = int(CENTRE_PAY_RATE * 100)


def _parse_created_at(created_at: str) -> datetime:
    text = created_at.replace("Z", "+00:00")
    dt = datetime.fromisoformat(text)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def _format_created_at(created_at: str) -> str:
    local = _parse_created_at(created_at).astimezone(stats_timezone())
    return local.strftime("%d/%m/%Y")


def _sorted_payments(records: list[PaymentRecord]) -> list[PaymentRecord]:
    return sorted(records, key=lambda r: (r.created_at, r.id), reverse=True)


def _user_label(
    username: str | None,
    display_name: str | None,
    user_id: int,
) -> str:
    name = (display_name or "").strip()
    if name:
        return name
    if username:
        return f"@{username.lstrip('@')}"
    return str(user_id)


def starter_payout(record: PaymentRecord) -> float:
    return _starter_payout(record)


def finisher_payout(record: PaymentRecord) -> float:
    return _finisher_payout(record)


def centre_payout(record: PaymentRecord) -> float:
    return _centre_payout(record)


def _starter_payout(record: PaymentRecord) -> float:
    if record.starter_user_id is None:
        return 0.0
    return record.amount * STARTER_PAY_RATE


def _finisher_payout(record: PaymentRecord) -> float:
    return record.amount * FINISHER_PAY_RATE


def _centre_payout(record: PaymentRecord) -> float:
    return record.amount * CENTRE_PAY_RATE


def _payment_rows(records: list[PaymentRecord]) -> list[list]:
    rows: list[list] = []
    for record in records:
        starter = ""
        if record.starter_user_id is not None:
            starter = _user_label(
                record.starter_username,
                record.starter_display_name,
                record.starter_user_id,
            )
        starter_pay = _starter_payout(record)
        finisher_pay = _finisher_payout(record)
        centre_pay = _centre_payout(record)
        rows.append(
            [
                _format_amount(record.amount),
                _format_created_at(record.created_at),
                starter,
                _user_label(
                    record.finisher_username,
                    record.finisher_display_name,
                    record.finisher_user_id,
                ),
                record.card_last4 or "",
                (
                    "Pending"
                    if record.cleared is None
                    else ("Yes" if record.cleared else "No")
                ),
                _format_amount(starter_pay) if starter_pay else "",
                _format_amount(finisher_pay),
                _format_amount(centre_pay),
            ]
        )
    return rows


def _footer_note() -> str:
    tz = stats_timezone()
    label = getattr(tz, "key", "local")
    now_local = datetime.now(timezone.utc).astimezone(tz)
    stamp = now_local.strftime(f"{now_local.day} %b %Y, %H:%M")
    return f"Updated {stamp} ({label})"


def payment_sheet_footer_note() -> str:
    """Footer stamp for payment sheet images and Excel export."""
    return _footer_note()


def format_payment_sheet_updated_note() -> str:
    """Short footer matching the Excel sheet (e.g. Updated 18 Jun 2026,)."""
    now_local = datetime.now(timezone.utc).astimezone(stats_timezone())
    stamp = now_local.strftime(f"{now_local.day} %b %Y,")
    return f"Updated {stamp}"


def sorted_payment_records(records: list[PaymentRecord]) -> list[PaymentRecord]:
    return _sorted_payments(records)


def payment_sheet_totals_row(records: list[PaymentRecord]) -> list[str]:
    total = sum(record.amount for record in records)
    total_starter = sum(_starter_payout(record) for record in records)
    total_finisher = sum(_finisher_payout(record) for record in records)
    total_centre = sum(_centre_payout(record) for record in records)
    count_label = f"{len(records)} payment" + ("" if len(records) == 1 else "s")
    return [
        "TOTAL",
        _format_amount(total),
        count_label,
        "",
        "",
        "",
        _format_amount(total_starter),
        _format_amount(total_finisher),
        _format_amount(total_centre),
    ]


def payment_sheet_data_rows(records: list[PaymentRecord]) -> list[list[str]]:
    return [[str(cell) for cell in row] for row in _payment_rows(records)]


def _export_rows(records: list[PaymentRecord]) -> list[list]:
    rows: list[list] = [list(HEADERS)]
    rows.extend(_payment_rows(records))
    rows.append(payment_sheet_totals_row(records))
    rows.append(["", "", "", "", "", "", "", "", _footer_note()])
    return rows


def _is_file_locked_error(exc: BaseException) -> bool:
    if isinstance(exc, PermissionError):
        return True
    return isinstance(exc, OSError) and getattr(exc, "winerror", None) == 5


def _atomic_replace(tmp_path: Path, output_path: Path) -> None:
    tmp_path.replace(output_path)


def _local_file_locked_message(settings: Settings) -> str:
    from onedrive_cloud_sync import graph_app_configured

    lines = [
        "q1.xlsx is open in another app (e.g. MobiOffice or Excel).",
        "Close it and run /syncpayments again.",
    ]
    if not graph_app_configured(settings):
        lines.append(
            "Tip: after /excelwebauth, cloud updates work even while the file is open locally."
        )
    return "\n".join(lines)


def _clear_sheet(ws) -> None:
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)


def _cleared_row_fill(cleared: bool | None):
    from openpyxl.styles import PatternFill

    if cleared is True:
        return PatternFill(fill_type="solid", fgColor="C6EFCE")
    if cleared is None:
        return PatternFill(fill_type="solid", fgColor="FFEB9C")
    return PatternFill(fill_type="solid", fgColor="FFC7CE")


def _style_sheet(
    ws,
    row_count: int,
    *,
    records: list[PaymentRecord] | None = None,
) -> None:
    from openpyxl.styles import Font

    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    if row_count >= 2:
        for cell in ws[row_count - 1]:
            cell.font = bold

    if not records:
        return

    for index, record in enumerate(records):
        row_num = index + 2
        fill = _cleared_row_fill(record.cleared)
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col).fill = fill


def _fill_sheet(
    ws,
    rows: list[list],
    *,
    records: list[PaymentRecord] | None = None,
) -> None:
    _clear_sheet(ws)
    for row in rows:
        ws.append(row)
    _style_sheet(ws, len(rows), records=records)


def _open_workbook_for_export(output_path: Path, worksheet_name: str):
    from openpyxl import Workbook, load_workbook

    if output_path.exists():
        try:
            wb = load_workbook(output_path)
            if worksheet_name in wb.sheetnames:
                ws = wb[worksheet_name]
            else:
                ws = wb.create_sheet(worksheet_name)
            return wb, ws
        except Exception as exc:
            logger.warning(
                "Could not open %s (%s); creating a new workbook",
                output_path.name,
                exc,
            )

    wb = Workbook()
    ws = wb.active
    ws.title = worksheet_name
    return wb, ws


def _build_workbook_bytes(
    output_path: Path,
    rows: list[list],
    *,
    worksheet_name: str,
    records: list[PaymentRecord] | None = None,
) -> bytes:
    wb, ws = _open_workbook_for_export(output_path, worksheet_name)
    try:
        _fill_sheet(ws, rows, records=records)
        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    finally:
        wb.close()


def _write_bytes_atomically(output_path: Path, content: bytes) -> None:
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = output_path.with_suffix(output_path.suffix + ".tmp")
    tmp_path.write_bytes(content)
    deadline = time.monotonic() + LOCAL_WRITE_RETRY_SECONDS
    attempt = 0
    while True:
        try:
            _atomic_replace(tmp_path, output_path)
            return
        except (PermissionError, OSError) as exc:
            if not _is_file_locked_error(exc) or time.monotonic() >= deadline:
                if tmp_path.exists():
                    try:
                        tmp_path.unlink()
                    except OSError:
                        pass
                raise
            attempt += 1
            time.sleep(min(3.0, 0.5 * attempt))


def _fallback_export_path(output_path: Path) -> Path:
    return output_path.with_name(f"{output_path.stem}-sync{output_path.suffix}")


def _write_workbook(
    settings: Settings,
    output_path: Path,
    records: list[PaymentRecord],
    *,
    worksheet_name: str,
) -> tuple[str, str | None]:
    """Write payments for Excel on the web. Returns (mode, web_url)."""
    from onedrive_cloud_sync import (
        graph_configured,
        remember_excel_web_url,
        resolve_excel_web_url,
        upload_workbook_to_onedrive,
    )

    remember_excel_web_url(settings)

    rows = _export_rows(records)
    content = _build_workbook_bytes(
        output_path,
        rows,
        worksheet_name=worksheet_name,
        records=records,
    )

    local_ok = False
    local_error: BaseException | None = None
    try:
        _write_bytes_atomically(output_path, content)
        local_ok = True
    except (PermissionError, OSError) as exc:
        if _is_file_locked_error(exc):
            local_error = exc
            logger.warning("Local Excel file is locked: %s", output_path)
        else:
            raise

    web_url: str | None = None
    if graph_configured(settings):
        web_url = upload_workbook_to_onedrive(settings, content)
        if web_url:
            if local_ok:
                return "cloud", web_url
            return "cloud_only", web_url

    if local_ok:
        return "file", resolve_excel_web_url(settings)

    if local_error is not None:
        alt_path = _fallback_export_path(output_path)
        try:
            _write_bytes_atomically(alt_path, content)
            logger.info(
                "Wrote payments export to %s because %s is locked",
                alt_path.name,
                output_path.name,
            )
            return "alt_file", str(alt_path)
        except (PermissionError, OSError):
            raise local_error
    return "file", resolve_excel_web_url(settings)


def _verify_workbook(
    output_path: Path,
    *,
    worksheet_names: list[str],
    min_rows_by_sheet: dict[str, int],
) -> None:
    from openpyxl import load_workbook

    if not output_path.exists():
        raise OSError(f"Export file missing after save: {output_path}")
    wb = load_workbook(output_path, read_only=True)
    try:
        for worksheet_name in worksheet_names:
            if worksheet_name not in wb.sheetnames:
                raise OSError(f"Sheet {worksheet_name!r} missing in {output_path.name}")
            min_rows = min_rows_by_sheet.get(worksheet_name, 1)
            row_count = sum(1 for _ in wb[worksheet_name].iter_rows(values_only=True))
            if row_count < min_rows:
                raise OSError(
                    f"Sheet {worksheet_name!r} looks incomplete "
                    f"({row_count} rows, expected at least {min_rows})"
                )
    finally:
        wb.close()


def _payments_for_export(
    settings: Settings,
    *,
    all_payments: bool = False,
) -> list[PaymentRecord]:
    if all_payments:
        records = list_all_payments(settings.database_path)
    else:
        epoch = get_paidside_epoch(settings.database_path)
        if epoch is not None:
            records = list_payments_since(settings.database_path, since=epoch)
        else:
            records = list_all_payments(settings.database_path)
    return _sorted_payments(records)


def export_payments_excel(
    settings: Settings,
    *,
    all_payments: bool = False,
) -> tuple[bool, str]:
    """Rewrite the Q1 payments workbook with every payment row."""
    from onedrive_cloud_sync import graph_configured, resolve_excel_web_url

    output_path = _export_output_path(settings)
    if output_path is None:
        return False, "PAYMENTS_ONEDRIVE_PATH is not set in .env"

    records = _payments_for_export(settings, all_payments=all_payments)
    paidside_epoch = (
        None if all_payments else get_paidside_epoch(settings.database_path)
    )
    sheet = settings.payments_onedrive_worksheet
    expected_min_rows = len(records) + 3
    sheet_checks = {sheet: expected_min_rows}

    try:
        refresh_mode, web_url = _write_workbook(
            settings,
            output_path,
            records,
            worksheet_name=sheet,
        )
        verify_path = output_path
        if refresh_mode == "alt_file" and web_url:
            verify_path = Path(web_url)
        if refresh_mode != "cloud_only":
            _verify_workbook(
                verify_path,
                worksheet_names=[sheet],
                min_rows_by_sheet=sheet_checks,
            )
    except (PermissionError, OSError) as exc:
        if _is_file_locked_error(exc):
            logger.warning("Payment export blocked — %s is open elsewhere", output_path)
            return False, _local_file_locked_message(settings)
        logger.exception("Payment export failed for %s", output_path)
        return False, f"Export failed: {exc}"
    except Exception as exc:
        logger.exception("Payment export failed for %s", output_path)
        return False, f"Export failed: {exc}"

    total = sum(record.amount for record in records)
    logger.info(
        "Exported %d payment(s) to %s (sheet %s, mode=%s)",
        len(records),
        output_path,
        sheet,
        refresh_mode,
    )
    mode = (
        "paid-side mode (new outs only)"
        if paidside_epoch is not None
        else "all payments"
    )
    lines = [
        f"Updated “{sheet}” in {output_path.name} "
        f"({len(records)} payments, {_format_amount(total)} total · {mode})",
    ]
    if refresh_mode == "alt_file" and web_url:
        alt_name = Path(web_url).name
        lines[0] = (
            f"Updated “{sheet}” ({len(records)} payments, "
            f"{_format_amount(total)} total · {mode})"
        )
        lines.append(
            f"Saved as {alt_name} — {output_path.name} is open in MobiOffice/Excel."
        )
        lines.append(
            f"Close {output_path.name}, then replace it with {alt_name} "
            "or run /syncpayments again."
        )
    elif refresh_mode == "cloud":
        lines.append("Pushed to your OneDrive file — Excel on the web is updated.")
        lines.append("If the tab is already open, press F5 once to refresh.")
    elif refresh_mode == "cloud_only":
        lines.append("Pushed to Excel on the web (local copy not updated — file was open).")
        lines.append("Close MobiOffice/Excel and run /syncpayments to refresh q1.xlsx on this PC.")
        lines.append("If the browser tab is open, press F5 once to refresh.")
    elif not graph_configured(settings):
        lines.append(
            "Saved locally only. Run /excelwebauth once (same Microsoft account as "
            "OneDrive), then /syncpayments — updates go straight to Excel on the web."
        )
    else:
        lines.append("Saved locally — cloud push failed. Try /syncpayments again.")
    if web_url and str(web_url).startswith("http"):
        lines.append(f"Excel on the web: {web_url}")
    return True, "\n".join(lines)


def _excel_sync_lock() -> asyncio.Lock:
    global _EXCEL_SYNC_LOCK
    if _EXCEL_SYNC_LOCK is None:
        _EXCEL_SYNC_LOCK = asyncio.Lock()
    return _EXCEL_SYNC_LOCK


async def run_payments_excel_sync(
    settings: Settings,
    *,
    all_payments: bool = False,
) -> tuple[bool, str]:
    if _export_output_path(settings) is None:
        return False, "PAYMENTS_ONEDRIVE_PATH is not set in .env"
    async with _excel_sync_lock():
        return await asyncio.to_thread(
            export_payments_excel, settings, all_payments=all_payments
        )


async def excel_sync_with_timer(
    bot,
    *,
    chat_id: int,
    message_id: int,
    settings: Settings,
    all_payments: bool = False,
    show_full_detail: bool = True,
) -> tuple[bool, str]:
    """Edit a Telegram message with a countdown until the Excel export finishes."""
    from telegram.error import BadRequest

    started = time.monotonic()
    countdown_end = started + SYNC_ESTIMATE_SECONDS
    sync_task = asyncio.create_task(
        run_payments_excel_sync(settings, all_payments=all_payments),
        name="payments-excel-sync-timer",
    )

    async def _set_status(text: str) -> None:
        try:
            await bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=text,
            )
        except BadRequest as exc:
            if "message is not modified" not in str(exc).lower():
                logger.debug("Could not edit Excel timer message: %s", exc)

    while not sync_task.done():
        remaining = max(0, int(countdown_end - time.monotonic()) + 1)
        if remaining > 0:
            await _set_status(f"⏳ Updating Excel… ~{remaining}s remaining")
        else:
            elapsed = max(1, int(time.monotonic() - started))
            await _set_status(f"⏳ Updating Excel… still working ({elapsed}s)")
        await asyncio.sleep(1)

    try:
        ok, detail = await sync_task
    except Exception as exc:
        logger.exception("Excel sync task crashed")
        await _set_status(f"❌ Excel update failed\n\n{exc}")
        return False, str(exc)

    elapsed = time.monotonic() - started
    if ok:
        body = detail if show_full_detail else detail.splitlines()[0]
        await _set_status(f"✅ Excel updated in {elapsed:.1f}s\n\n{body}")
        logger.info("Excel sync with timer: %s", detail.splitlines()[0])
    else:
        await _set_status(f"❌ Excel update failed ({elapsed:.1f}s)\n\n{detail}")
        logger.error("Excel sync with timer failed: %s", detail)
    return ok, detail


async def _run_payments_excel_sync(settings: Settings) -> None:
    ok, detail = await run_payments_excel_sync(settings)
    if ok:
        logger.info("Background Excel sync: %s", detail.splitlines()[0])
    else:
        logger.error("Background Excel sync failed: %s", detail)


def schedule_payments_excel_sync(settings: Settings) -> None:
    if _export_output_path(settings) is None:
        return
    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        return
    loop.create_task(
        _run_payments_excel_sync(settings),
        name="payments-excel-sync",
    )
